import requests
import json
import openpyxl
import time


def find_city_code(name):
    counter = 2
    while counter<10:
        if citycodes.cell(row=counter,column=1).value == name:
            return citycodes.cell(row=counter,column=2).value
        counter+=1
    return None

def CtoF(c):
    F = ((c*9)/5)+32
    return F

parameters = {'appid':'c0766fb14c2e6c9244346f7cd2a256ec','id':''}

#Intermediate file to write the updated Temperature - Linked to the final file.
wb = openpyxl.load_workbook(filename='Meanwhile.xlsm', read_only=False, keep_vba=True)
weatherSheet = wb['LiveData'] #Sheet with city names to be displayed
citycodes = wb['CityCodes'] #Sheet with database of city codes

for i in range(2,7):
    CName = weatherSheet.cell(row=i, column=1).value
    status = weatherSheet.cell(row=i, column=5).value
    CityID = find_city_code(CName)
    unit = weatherSheet.cell(row=i, column=4).value
    parameters['id']=CityID

    #Get Info using API
    response = requests.get("http://api.openweathermap.org/data/2.5/weather?", params=parameters).json()
    temp = response["main"]['temp'] - 273.15 
    humidity = response["main"]['humidity']
    if unit == "F":
        temp = CtoF(temp)

    #Export the data to respective cells
    weatherSheet.cell(row=i, column=2).value = temp
    weatherSheet.cell(row=i, column=3).value = humidity
else:
    wb.save('Meanwhile.xlsm')

while True:
    print("Updating every second...")
    time.sleep(1)
    for i in range(2,7):
        CName = weatherSheet.cell(row=i, column=1).value
        status = weatherSheet.cell(row=i, column=5).value
        CityID = find_city_code(CName)
        unit = weatherSheet.cell(row=i, column=4).value
        if status:
            parameters['id']=CityID
            response = requests.get("http://api.openweathermap.org/data/2.5/weather?", params=parameters).json()
            temp = response["main"]['temp'] - 273.15 
            humidity = response["main"]['humidity']
            if unit == "F":
                temp = CtoF(temp)
            weatherSheet.cell(row=i, column=2).value = temp
            weatherSheet.cell(row=i, column=3).value = humidity

wb.save('Meanwhile.xlsm')